from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.audit_logger import mapping_actions_to_dataframe, validation_issues_to_dataframe
from app.models import MappingAction, ValidationResult
from app.utils import EXCEL_ROW_COLUMN


def generate_cleaned_template(
    original_template_path: str,
    mapped_template_data: dict[str, pd.DataFrame],
    mapping_actions: list[MappingAction],
    validation_result: ValidationResult,
    output_path: str,
) -> str:
    """Generate an auditable cleaned workbook from mapped template data."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelFile(original_template_path) as original_workbook:
        sheet_order = list(original_workbook.sheet_names)

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            written_sheets: set[str] = set()
            for sheet_name in sheet_order:
                if sheet_name in mapped_template_data:
                    _public_dataframe(mapped_template_data[sheet_name]).to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False,
                    )
                else:
                    original_df = pd.read_excel(original_workbook, sheet_name=sheet_name, dtype=object)
                    original_df.to_excel(writer, sheet_name=sheet_name, index=False)
                written_sheets.add(sheet_name)

            for sheet_name, df in mapped_template_data.items():
                if sheet_name in written_sheets:
                    continue
                _public_dataframe(df).to_excel(writer, sheet_name=sheet_name, index=False)

            mapping_actions_to_dataframe(mapping_actions).to_excel(
                writer,
                sheet_name="Mapping Audit",
                index=False,
            )
            validation_issues_to_dataframe(validation_result.issues).to_excel(
                writer,
                sheet_name="Validation Issues",
                index=False,
            )
            _add_mapping_comments(writer.book, mapping_actions)
            _format_workbook(writer.book)

    return str(output)


def _public_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    public_columns = [column for column in df.columns if column != EXCEL_ROW_COLUMN]
    return df.loc[:, public_columns].copy()


def _add_mapping_comments(workbook: Any, mapping_actions: list[MappingAction]) -> None:
    for action in mapping_actions:
        if action.status != "MAPPED":
            continue
        if action.sheet_name not in workbook.sheetnames or action.row_number <= 1:
            continue

        sheet = workbook[action.sheet_name]
        header_columns = {cell.value: cell.column for cell in sheet[1]}
        column_index = header_columns.get(action.field_name)
        if not column_index:
            continue
        if action.row_number > sheet.max_row:
            continue

        cell = sheet.cell(row=action.row_number, column=column_index)
        cell.comment = Comment(
            (
                f"Mapping: {action.mapping_name}\n"
                f"Original: {action.original_value}\n"
                f"Mapped: {action.mapped_value}\n"
                f"Status: {action.status}"
            ),
            "SAP Migration Prevalidator",
        )


def _format_workbook(workbook: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    audit_fill = PatternFill("solid", fgColor="DDEBF7")
    issue_fill = PatternFill("solid", fgColor="FCE4D6")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        if sheet.title == "Mapping Audit":
            for row in sheet.iter_rows(min_row=2):
                status = str(row[7].value or "")
                if status == "MAPPED":
                    for cell in row:
                        cell.fill = audit_fill
        elif sheet.title == "Validation Issues":
            for row in sheet.iter_rows(min_row=2):
                severity = str(row[5].value or "")
                if severity == "ERROR":
                    for cell in row:
                        cell.fill = issue_fill

        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            column_letter = get_column_letter(column_index)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)
