from __future__ import annotations

from collections.abc import Sequence
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import ValidationIssue, ValidationResult


DETAIL_COLUMNS = [
    "Object",
    "Sheet",
    "Row",
    "Field",
    "Value",
    "Rule Type",
    "Message",
    "Suggested Fix",
]
LLM_EXPLANATION_COLUMN = "LLM Explanation"


def generate_excel_report(
    result: ValidationResult,
    output_path: str | Path,
    llm_explanations: Sequence[str] | None = None,
) -> str:
    report_bytes = generate_excel_report_bytes(result, llm_explanations=llm_explanations)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(report_bytes)
    return str(output)


def generate_excel_report_bytes(
    result: ValidationResult,
    llm_explanations: Sequence[str] | None = None,
) -> bytes:
    explanations_by_id = _explanations_by_id(result, llm_explanations)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _summary_df(result).to_excel(writer, sheet_name="Summary", index=False)
        _issues_df(
            [issue for issue in result.issues if issue.severity == "ERROR"],
            explanations_by_id,
        ).to_excel(
            writer,
            sheet_name="Error Details",
            index=False,
        )
        _issues_df(
            [issue for issue in result.issues if issue.severity == "WARNING"],
            explanations_by_id,
        ).to_excel(
            writer,
            sheet_name="Warning Details",
            index=False,
        )
        _issues_df(
            [issue for issue in result.issues if issue.suggested_fix],
            explanations_by_id,
        ).to_excel(
            writer,
            sheet_name="Suggested Fixes",
            index=False,
        )
        _format_workbook(writer.book)
    return output.getvalue()


def _summary_df(result: ValidationResult) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Object": result.object_name,
                "Total Rows": result.total_rows,
                "Error Count": result.error_count,
                "Warning Count": result.warning_count,
                "Status": result.status,
            }
        ]
    )


def _issues_df(
    issues: list[ValidationIssue],
    explanations_by_id: dict[int, str] | None = None,
) -> pd.DataFrame:
    rows = [
        {
            "Object": issue.object_name,
            "Sheet": issue.sheet_name,
            "Row": issue.row_number,
            "Field": issue.field_name,
            "Value": issue.value,
            "Rule Type": issue.rule_type,
            "Message": issue.message,
            "Suggested Fix": issue.suggested_fix,
            **(
                {LLM_EXPLANATION_COLUMN: explanations_by_id.get(id(issue), "")}
                if explanations_by_id is not None
                else {}
            ),
        }
        for issue in issues
    ]
    columns = DETAIL_COLUMNS.copy()
    if explanations_by_id is not None:
        columns.append(LLM_EXPLANATION_COLUMN)
    return pd.DataFrame(rows, columns=columns)


def _explanations_by_id(
    result: ValidationResult,
    llm_explanations: Sequence[str] | None,
) -> dict[int, str] | None:
    if llm_explanations is None:
        return None
    return {
        id(issue): str(llm_explanations[index])
        for index, issue in enumerate(result.issues)
        if index < len(llm_explanations)
    }


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            column_letter = get_column_letter(column_index)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
