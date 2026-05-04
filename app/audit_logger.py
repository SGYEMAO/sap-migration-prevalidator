from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import MappingAction, ValidationIssue, ValidationResult


MAPPING_AUDIT_COLUMNS = [
    "Object",
    "Sheet",
    "Row",
    "Field",
    "Original Value",
    "Mapped Value",
    "Mapping Name",
    "Status",
    "Confidence",
    "Message",
]

VALIDATION_ISSUE_COLUMNS = [
    "Object",
    "Sheet",
    "Row",
    "Field",
    "Value",
    "Severity",
    "Rule Type",
    "Message",
    "Suggested Fix",
]


def mapping_actions_to_dataframe(mapping_actions: list[MappingAction]) -> pd.DataFrame:
    rows = [
        {
            "Object": action.object_name,
            "Sheet": action.sheet_name,
            "Row": action.row_number,
            "Field": action.field_name,
            "Original Value": action.original_value,
            "Mapped Value": action.mapped_value,
            "Mapping Name": action.mapping_name,
            "Status": action.status,
            "Confidence": action.confidence,
            "Message": action.message,
        }
        for action in mapping_actions
    ]
    return pd.DataFrame(rows, columns=MAPPING_AUDIT_COLUMNS)


def validation_issues_to_dataframe(issues: list[ValidationIssue]) -> pd.DataFrame:
    rows = [
        {
            "Object": issue.object_name,
            "Sheet": issue.sheet_name,
            "Row": issue.row_number,
            "Field": issue.field_name,
            "Value": issue.value,
            "Severity": issue.severity,
            "Rule Type": issue.rule_type,
            "Message": issue.message,
            "Suggested Fix": issue.suggested_fix,
        }
        for issue in issues
    ]
    return pd.DataFrame(rows, columns=VALIDATION_ISSUE_COLUMNS)


def generate_mapping_audit_report(
    mapping_actions: list[MappingAction],
    output_path: str | Path,
) -> str:
    report_bytes = generate_mapping_audit_report_bytes(mapping_actions)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(report_bytes)
    return str(output)


def generate_mapping_audit_report_bytes(mapping_actions: list[MappingAction]) -> bytes:
    output = BytesIO()
    audit_df = mapping_actions_to_dataframe(mapping_actions)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _mapping_summary_df(mapping_actions).to_excel(writer, sheet_name="Summary", index=False)
        audit_df.to_excel(writer, sheet_name="Mapping Audit", index=False)
        _format_workbook(writer.book)
    return output.getvalue()


def _mapping_summary_df(mapping_actions: list[MappingAction]) -> pd.DataFrame:
    counts = {
        "MAPPED": 0,
        "UNMAPPED": 0,
        "AMBIGUOUS": 0,
        "UNCHANGED": 0,
    }
    for action in mapping_actions:
        counts[action.status] += 1
    return pd.DataFrame(
        [
            {
                "Mapped": counts["MAPPED"],
                "Unmapped": counts["UNMAPPED"],
                "Ambiguous": counts["AMBIGUOUS"],
                "Unchanged": counts["UNCHANGED"],
                "Total Actions": len(mapping_actions),
            }
        ]
    )


def add_validation_issues_sheet(
    writer: pd.ExcelWriter,
    validation_result: ValidationResult,
    sheet_name: str = "Validation Issues",
) -> None:
    validation_issues_to_dataframe(validation_result.issues).to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
    )


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    status_fills = {
        "MAPPED": PatternFill("solid", fgColor="E2F0D9"),
        "UNMAPPED": PatternFill("solid", fgColor="FCE4D6"),
        "AMBIGUOUS": PatternFill("solid", fgColor="FFF2CC"),
        "UNCHANGED": PatternFill("solid", fgColor="DDEBF7"),
    }

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        header_by_name = {cell.value: cell.column for cell in sheet[1]}
        status_column = header_by_name.get("Status")
        if status_column:
            for row in sheet.iter_rows(min_row=2, min_col=status_column, max_col=status_column):
                status_cell = row[0]
                fill = status_fills.get(str(status_cell.value))
                if fill:
                    status_cell.fill = fill

        for column_cells in sheet.columns:
            column_index = column_cells[0].column
            column_letter = get_column_letter(column_index)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)
