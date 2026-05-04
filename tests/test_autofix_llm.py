from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.autofix_engine import generate_cleaned_template
from app.config_loader import load_config_files
from app.data_parser import load_template_excel_with_profile
from app.local_llm import explain_issue_with_local_llm
from app.mapping_engine import apply_mappings, load_mapping_files
from app.models import ValidationIssue
from app.profile_loader import load_profile, required_mapping_filenames
from app.rule_engine import run_validation


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def test_autofix_generates_cleaned_template(tmp_path: Path) -> None:
    profile = load_profile("MATERIAL_MASTER", PROJECT_ROOT / "profiles")
    template_path = PROJECT_ROOT / "sample_templates" / "material_master_legacy_values_sample.xlsx"
    template_data = load_template_excel_with_profile(template_path, profile)
    mapping_sources = {
        filename: PROJECT_ROOT / "mappings" / "MATERIAL_MASTER" / filename
        for filename in required_mapping_filenames(profile).values()
    }
    mappings = load_mapping_files(profile, mapping_sources)
    mapped_data, mapping_actions = apply_mappings(
        "MATERIAL_MASTER",
        profile,
        template_data,
        mappings,
    )
    config_data = load_config_files(profile, fallback_dir=PROJECT_ROOT / "config_samples")
    validation_result = run_validation("MATERIAL_MASTER", profile, mapped_data, config_data)

    output_path = tmp_path / "cleaned_template.xlsx"
    cleaned_path = generate_cleaned_template(
        str(template_path),
        mapped_data,
        mapping_actions,
        validation_result,
        str(output_path),
    )

    assert Path(cleaned_path).exists()
    basic_data = pd.read_excel(cleaned_path, sheet_name="Basic Data", dtype=object)
    plant_data = pd.read_excel(cleaned_path, sheet_name="Plant Data", dtype=object)
    audit = pd.read_excel(cleaned_path, sheet_name="Mapping Audit", dtype=object)

    assert "__excel_row_number" not in basic_data.columns
    assert basic_data["MaterialType"].tolist() == ["FERT", "ROH", "ROH"]
    assert basic_data["BaseUoM"].tolist() == ["EA", "PC", "BADUOM"]
    assert plant_data["Plant"].tolist() == ["1000", "SG99", "1100"]
    assert "MAPPED" in set(audit["Status"])
    assert "UNMAPPED" in set(audit["Status"])

    workbook = load_workbook(cleaned_path)
    assert "Validation Issues" in workbook.sheetnames
    assert workbook["Basic Data"]["C2"].comment is not None


def test_llm_fallback_when_ollama_unavailable(monkeypatch) -> None:
    def raise_connection_error(*args, **kwargs):
        raise OSError("Ollama is not running")

    monkeypatch.setattr("urllib.request.urlopen", raise_connection_error)
    issue = ValidationIssue(
        object_name="MATERIAL_MASTER",
        sheet_name="Basic Data",
        row_number=2,
        field_name="BaseUoM",
        value="BADUOM",
        severity="ERROR",
        rule_type="CONFIG_CHECK",
        message="BaseUoM 'BADUOM' does not exist in config source 'uom'.",
        suggested_fix="Use a value listed in uom.xlsx.",
    )

    explanation = explain_issue_with_local_llm(issue, context={}, model_name="llama3.1:8b")
    parsed = json.loads(explanation)

    assert parsed["risk_level"] == "HIGH"
    assert "BADUOM" in parsed["business_explanation"]
    assert parsed["recommended_action"] == "Use a value listed in uom.xlsx."
